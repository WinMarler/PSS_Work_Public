import os
import re
import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import Alignment

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def export_html_to_excel():
    """将table_html文件夹下的所有HTML文件导出到Excel"""
    folder_path = os.path.dirname(__file__)
    output_path = os.path.join(folder_path, "html_tables.xlsx")
    
    data = []
    
    # 遍历所有HTML文件
    for filename in os.listdir(folder_path):
        if filename.startswith("table_seq") and filename.endswith(".html"):
            try:
                # 提取序号
                seq = re.search(r'table_seq(\d+)\.html', filename).group(1)
                
                # 读取HTML内容
                file_path = os.path.join(folder_path, filename)
                with open(file_path, 'r', encoding='utf-8') as f:
                    html_content = f.read()
                
                data.append({
                    "Seq": int(seq),
                    "Table_code": html_content
                })
                logger.info(f"成功处理: {filename}")
                
            except Exception as e:
                logger.error(f"处理文件 {filename} 失败: {str(e)}")
    
    # 生成Excel文件
    if data:
        df = pd.DataFrame(data).sort_values('Seq')
        df.to_excel(output_path, index=False, engine='openpyxl')
        
        # 自动调整列宽和行高
        wb = load_workbook(output_path)
        ws = wb.active
        
        # 设置列宽（A列15，B列120）
        ws.column_dimensions['A'].width = 15  # Seq列
        ws.column_dimensions['B'].width = 120  # HTML代码列
        
        # 设置自动换行和垂直对齐
        for row in ws.iter_rows(min_row=2, max_col=2):
            cell = row[1]  # B列
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
        # 自动调整行高（根据内容）
        for row in ws.iter_rows(min_row=2):
            max_lines = max([len(str(cell.value).split('\n')) for cell in row])
            ws.row_dimensions[row[0].row].height = min(100, max_lines * 15)
        
        wb.save(output_path)
        logger.info(f"Excel文件已生成并自动调整格式: {output_path}")
    else:
        logger.warning("未找到有效的HTML文件")

if __name__ == "__main__":
    export_html_to_excel()